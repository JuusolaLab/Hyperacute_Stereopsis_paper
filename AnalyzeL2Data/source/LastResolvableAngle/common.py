import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import re
import xlsxwriter
import openpyxl
import os
from imageio import volread
import tifffile
import xml.dom.minidom
import cv2
import xlwings as xw
import time
import math
#constants
dist_screen_eyes = 38
bar90t50 = 2.67
bar90t20 = 1.11
bart100 = 9 #(mm)since new scaling, this is the width of a bar of 100px
time_course_calcium = .1 #s
speed_90bars30 = .694 #mm/s
screens_ratio = 2

#PATHS:
#drive = '//smallbrains-nas2/keivan/'
drive = 'E:/'
input_path = drive+'Data/' 
path_to_stimuli = input_path+'Track ball/Stimuli/'
output_path = input_path+'Plot/' 

def time_offset(wrong_time,stim_time,t_stop,nbf,frame_rate):
	real_rate = t_stop/stim_time*frame_rate
	real_time = nbf/real_rate
	return wrong_time-real_time
print('210126000000 offset: ',int(100*time_offset(53.13,45,38.6,4000,100))/100,'s')
print('210127000001 offset: ',int(100*time_offset(51.85,40,17.7,750,37.5))/100,'s')
print('210201000000 offset: ',int(100*time_offset(54.35,40,16.7,3000,150))/100,'s')

def stim_contin(len):
	stim = np.zeros(len)
	stim[int(len/8):int(3*len/8)] = np.arange(int(len/8),int(3*len/8),1)
	stim[int(3*len/8):int(5*len/8)] = [int(3*len/8)]*(int(5*len/8)-int(3*len/8))
	stim[int(5*len/8):int(7*len/8)] = [int(3*len/8)-x for x in np.arange(int(len/8),int(3*len/8),1)]
	return(stim)
	
def peaks_index(x,N):
	y = np.convolve(x, np.ones((N,))/N, mode='same')
	peakind = signal.find_peaks_cwt(x, np.arange(5,20))
	print(peakind)
	return peakind

def read_tiff_timevideo(file):
	bonjour = tifffile.TiffFile(file)
	num_string = bonjour.ome_metadata.split('Length_2=')[1]
	return float(num_string.split('"')[1])

def read_tiff_pixelsize(file):
	bonjour = tifffile.TiffFile(file)
	num_stringX = bonjour.ome_metadata.split('PhysicalSizeX=')[1]
	num_stringY = bonjour.ome_metadata.split('PhysicalSizeY=')[1]
	return (float(num_stringX.split('"')[1])+float(num_stringY.split('"')[1]))/2

def fitDoubleGaussian(xaxis,trace,dt):
	tmax = xaxis[np.argmax(trace)]
	t1 = tmax-dt
	t2 = tmax+dt
	gaussian = lambda x, a, b, c : a * np.exp((-1)*(x-b)**2/(2*c**2)) 
	double_gaussian1 = lambda x,a,c,t: gaussian(x,max(trace)-np.median(trace),t,c) + gaussian(x,a,t-dt,c) + np.median(trace)
	double_gaussian2 = lambda x,a,c,t: gaussian(x,max(trace)-np.median(trace),t,c) + gaussian(x,a,t+dt,c) + np.median(trace)
	bounds = ([0.1*max(trace),0.5,tmax-2],[max(trace),2,tmax+2])
	p0=(0.5*max(trace),1,tmax)
	popt1, pcov1 = curve_fit(double_gaussian1, xaxis, trace, p0=p0, bounds = bounds)
	popt2, pcov2 = curve_fit(double_gaussian2, xaxis, trace, p0=p0, bounds = bounds)
	print(popt1,popt2)
	res1 = resolvability(xaxis,trace,[popt1[2],popt1[2]-dt])
	res2 = resolvability(xaxis,trace,[popt2[2],popt2[2]+dt])
	print(res1['resol'],res2['resol'])
	if res1['resol']>res2['resol']:
		popt = popt1
		double_gaussian = double_gaussian1
		res = res1
	else:
		popt = popt2
		double_gaussian = double_gaussian2
		res = res2
	print('optimal parameters for gaussian fit: ',popt)
	
	gp = gradient_pmax(xaxis,trace,dt)
	print(gp['resol'])
	#if gp['resol']>res['resol']:
		#return {'trace':np.gradient(trace),'resol':gp['resol'], 'pmax1':gp['pmax1'],'pmax2':gp['pmax2'],'pmin':gp['pmin']}

	return {'trace':double_gaussian(xaxis, *popt),'resol':res['resol'], 'pmax1':res['pmax1'],'pmax2':res['pmax2'],'pmin':res['pmin']}

def parse_get_time(file,video):
	res = []
	with open(file, 'r') as f:
		data = f.readlines()
	for line in data[:-1]:
		numbers = line.split()
		if (numbers[0]==video):
			return float(numbers[1])
	return 0
	
def first_rise(trace,thresh,x1,x2,nt,narrowing = False):
	nonoise_trace = trace[nt:]
	if narrowing:
		indexes = np.arange(len(nonoise_trace)-x2,len(nonoise_trace)-x1)
	else:
		indexes = np.arange(x1,x2)
	minbase = np.mean(nonoise_trace[indexes])+thresh*np.std(nonoise_trace[indexes])
	return nt+np.where(nonoise_trace<=minbase)[0][-1*(not narrowing)]


def ntsigma(traces,nt):
	stds = np.zeros(len(traces))
	for i in range(0,len(traces)):
		stds[i] = np.std(traces[i][0:nt])
	avg = np.mean(stds)
	std = np.std(stds)
	return {'avg':avg,'std':std}

def ntSNR(traces,nt):
	SNRs = np.zeros(len(traces))
	for i in range(0,len(traces)):
		SNRs[i] = (np.mean(traces[i][0:nt])/np.std(traces[i][0:nt]))**2
	avg = np.mean(SNRs)
	std = np.std(SNRs)
	return {'avg':avg,'std':std}
#print(rfmatrix([479,458],[512,384],270,1000))

def italic_frame(m,a):
	print('size of image:',m.shape)
	m2 = m.copy()
	for j in range(0,m.shape[0]):
		dis = int(np.tan(a/180*3.14)*j)
		#print(dis)
		for i in range(dis,m.shape[1]):
			m2[j,i] = m[j,i-dis]
		for i in range(0,dis):
				m2[j,i] = 0
	return m2[:,dis:m.shape[1]]
def frame_scan(m,w,dt):
	#cv2.imshow('frame0',m)
	out = cv2.VideoWriter(output_path+'dyna_grating.avi',cv2.VideoWriter_fourcc('M','J','P','G'), 30, (w,m.shape[0]),False)
	for t in range(0,m.shape[1]-w,dt):
		frame = m[:,t:t+w]
		out.write(frame)
		#cv2.imshow('frame',frame)
		if cv2.waitKey(1) & 0xFF == ord('q'):
		  break
	out.release()
	cv2.destroyAllWindows() 
def mat_to_video(m,fps,name):
	out = cv2.VideoWriter(name,cv2.VideoWriter_fourcc('M','J','P','G'), fps, (m.shape[2],m.shape[1]),False)
	max = m.max()
	for t in range(0,m.shape[0]):
		out.write((m[t]/max*2**8).astype('uint8'))
	out.release()
	cv2.destroyAllWindows() 

def trunc(x,n):
	return int(10**n*x)/10**n

def angle_to_180(x):
	while x>180:
		x = x-180
	while x<0:
		x = x+180
	return x

def angle180_to_90(x):
	if x>90:
		x = 180-x
	return x

def xy_to_angle(x,y):
	if x == 0:
		if y <0: return -90
		else: return 90
	else:
		a = np.arctan(y/x)
		if x<0: a=3.14+a
	return 180/3.14*a
	
def ltd_times(frames,x,opposite = False):
	times = []
	for i in range(1,len(frames)):
		if opposite:
			if frames[i][1,x]>frames[i-1][1,x]:
				times.append(i)
		else:
			if frames[i][1,x]<frames[i-1][1,x]:
				times.append(i)
	times = np.asarray(times)
	return times

def ltd_times_nev(frames,x):
	times = []
	for i in range(1,frames.shape[2]):
		if frames[1,x,i]<frames[1,x,i-1]:
			times.append(i)
	times = np.asarray(times)
	return times

def ltd_peaks(trace,ltd_indexes,m=.1,p=.1):
	peaks = []
	if len(ltd_indexes)<3:
		firstid = 0
	else:
		firstid = 1
	t0 = ltd_indexes[firstid]
	dt = ltd_indexes[firstid+1]-ltd_indexes[firstid]
	tmax = t0-int(dt/2)+np.argmax(trace[(t0-int(dt/2)):(t0+int(dt/2))])
	peaks.append(tmax)
	while tmax<len(trace)-2 and len(peaks)<len(ltd_indexes):
		k = max(1,np.argmin(np.abs(tmax-ltd_indexes)))
		dt = ltd_indexes[k]-ltd_indexes[k-1]
		if len(trace[(tmax+dt-int(m*dt)):(tmax+dt+int(m*dt)+1)])==0:
			break
		tmax = min(len(trace)-2,tmax+dt-int(m*dt)+np.argmax(trace[(tmax+dt-int(m*dt)):(tmax+dt+int(m*dt)+1)]))
		#check possible extension if not on a peak yet
		if p>0:
			if trace[tmax-1]>trace[tmax]:
				shift=0
				while trace[tmax-1]>trace[tmax] and shift<p*dt and tmax>0:
					tmax -= 1
					shift += 1
			elif trace[tmax+1]>trace[tmax]:
				shift=0
				while trace[tmax+1]>trace[tmax] and shift<p*dt and tmax<len(trace)-2:
					tmax += 1
					shift += 1
		peaks.append(tmax)
	return(np.asarray(peaks))

def align_times_peaks_rf(trace,frames,x1,x2,step):
	s = 0
	times = []
	minlen = len(trace)
	for x in range(x1,x2,step):
		times.append(ltd_times(frames,x))
		minlen = min(minlen,len(times[-1]))
	i = 0
	for time in times:
		s1 = sum(trace[time[0:30]]) 
		print(s1)
		if s1>s:
			s = s1
			newtimes = time
			x = x1+i*step
		i+=1
	print("Receptive field at ",x)
	return newtimes
	
def align_times_peaks(trace,times,min,max,step,time_video,n):
	s = 0
	timescut = times
	i = 0
	for t in timescut:
		if t+min<0 or t+max+int(time_course_calcium/time_video*n)>=len(trace):
			timescut = np.delete(timescut,i,0)
		else: 
			i+=1
	for k in range(min,max,step):
		s1 = sum(np.gradient(trace)[timescut[0:5]+k])
		if s1>s:
			s = s1
			newtimes = timescut+k+int(time_course_calcium/time_video*n)
	return newtimes



def real_angle(dir,ratio):
	if dir>3.14/2 and dir<3*3.14/2:
		return np.arctan(1/ratio*np.tan(dir))+3.14
	return np.arctan(1/ratio*np.tan(dir))
	

def real_dist(d,dir1,dir2,ratio):
	return ratio*d*np.cos(dir1)/np.cos(dir2)*np.cos(dir1-dir2)
	
def real_roi_pos(pts,cpx,length):
	pts2 = []
	for pt in pts:
		pts2.append(cpx*length+pt)
	return pts2

def pixel_to_angle_bars(p,dir):
	#for new data: dir = 0
	dir = 0
	dir1 = 3.14/180*dir
	dir2 = real_angle(dir1,screens_ratio)
	d = bar90t20+(p-20)/(50-20)*(bar90t50-bar90t20)
	d2 = real_dist(d,dir1,dir2,screens_ratio)
	theta = d2/dist_screen_eyes*180/3.14
	return theta

def pixel_to_angle_bars2(p):
	d = bart100/100*p
	theta = np.arctan(d/dist_screen_eyes)*180/3.14
	return int(100*theta)/100
print(pixel_to_angle_bars2(1),' degrees per pixel (small distances) (should be 0.13°)')
	
def pixel_to_angle_bars_nev(p):
		return 0.0417*p

	
def type_of_image(image):
	if(len(image.shape)<3):
		  return 'gray'
	elif len(image.shape)==3:
		  return 'color'
	else:
		  return 'others'
def gac(image):
	if type_of_image(image) == 'color':
		gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
		color = image
	else:
		gray = image.copy()
		color = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
	return {'gray':gray, 'color':color}
	
def draw_ROI_colours(frame,rois,colours_indexes,plot=False):
	i = 0
	output = frame.copy()
	for roi in rois:
		cv2.rectangle(output, (roi[0], roi[1]), (roi[2], roi[3]), colorsRGB2[colours_indexes[i]], 2)
		i+=1
	if plot:
		cv2.imshow("output", output)
		cv2.waitKey(0)
	return output

def draw_ROI_SRAcolours(frame,rois,colours_indexes,SRAs,plot=False,angle_scale = [2.5,5.5]):
	i = 0
	output = frame.copy()
	for roi in rois:
		if SRAs[colours_indexes[i]] is None:
			cv2.rectangle(output, (roi[0], roi[1]), (roi[2], roi[3]), (0, 0, 0), 2)
		else:
			blue_colour = int(255*(SRAs[colours_indexes[i]]-angle_scale[0])/(angle_scale[1]-angle_scale[0]))
			cv2.rectangle(output, (roi[0], roi[1]), (roi[2], roi[3]), (blue_colour, 0, 255-blue_colour), 2)
		i+=1
	if plot:
		cv2.imshow("output", output)
		cv2.waitKey(0)
	return output
	
def rfpos_frame(frame,rois,colours_indexes,rfpos,plot=False):
	i = 0
	output = np.full(frame.shape,None).astype('float')
	for roi in rois:
		if rfpos[colours_indexes[i]] is not None:
			output[roi[1]:roi[3],roi[0]:roi[2]] = rfpos[colours_indexes[i]]
		i+=1
	if plot:
		cv2.imshow("rfpos", mat_to_frame(output,bits = 8))
	return output

def rfpos_by_cluster(rfpos0,rfpos90,rois,labels):
	nclusters = max(labels)+1
	rfpos0_by_cluster = [None]*nclusters
	rfpos90_by_cluster = [None]*nclusters
	for i in range(len(labels)):
		mean0 = np.nanmean(rfpos0[rois[i][1]:rois[i][3],rois[i][0]:rois[i][2]])
		if not np.isnan(mean0):
			if rfpos0_by_cluster[labels[i]] is None:
				rfpos0_by_cluster[labels[i]] = mean0
			else :
				rfpos0_by_cluster[labels[i]] += mean0
		mean90 = np.nanmean(rfpos90[rois[i][1]:rois[i][3],rois[i][0]:rois[i][2]])
		if not np.isnan(mean90):
			if rfpos90_by_cluster[labels[i]] is None:
				rfpos90_by_cluster[labels[i]] = mean90
			else:
				rfpos90_by_cluster[labels[i]] += mean90
	rfpos0_by_cluster = np.asarray(rfpos0_by_cluster).astype('float')/nclusters
	rfpos90_by_cluster = np.asarray(rfpos90_by_cluster).astype('float')/nclusters
	return {'rfpos0':rfpos0_by_cluster,'rfpos90':rfpos90_by_cluster}
	
def com_by_cluster(rois,labels):
	from sklearn.decomposition import PCA
	nclusters = max(labels)+1
	meanx = np.zeros(nclusters)
	meany = np.zeros(nclusters)
	sizes = np.zeros(nclusters)
	evectors = []
	evalues = []
	rois = np.asarray(rois)
	for i in range(nclusters):
		points = rois[np.where(labels==i)]
		sizes[i] = len(points)
		meansx = (points[:,3]+points[:,1])/2
		meansy = (points[:,2]+points[:,0])/2
		meanx[i] = np.mean(meansx)
		meany[i] = np.mean(meansy)
		if sizes[i] <= 1:
			evectors.append([[0,0],[0,0]])
			evalues.append([0,0])
		else:
			pca = PCA(n_components=2)
			xy_train = np.vstack((meansx-meanx[i],meansy-meany[i])).T
			pca.fit(xy_train)
			evectors.append(pca.components_)
			evalues.append(pca.explained_variance_)
	return {'meanx':meanx,'meany':meany,'evectors':evectors,'evalues':evalues,'sizes':sizes}
	
def getbits(x):
	num = re.findall(r'\d+', str(x.dtype)) 
	return int(num[0])

def scale(frame,bits = None):
	if bits == None:
		bits = getbits(frame)
	rep = (frame - frame.min())/(frame.max() - frame.min())*2**bits
	return rep.astype('uint'+str(bits))
	
def to8(frame):
	bits = getbits(frame)
	return (frame/2**(bits-8)).astype('uint8')

def read_tiffstacks(file,crop=None):
	print("starting reading...")
	frames = volread(file)
	print("finished reading...")
	if crop is None:
		return frames
	else:
		t0=int(frames.shape[0]*crop[0])
		t1=int(frames.shape[0]*crop[1])
		x0=int(frames.shape[1]*crop[2])
		x1=int(frames.shape[1]*crop[3])
		y0=int(frames.shape[2]*crop[4])
		y1=int(frames.shape[2]*crop[5])
		return frames[t0:t1,x0:x1,y0:y1]

def bkg_pos(size,pos,shape):
	return [(shape[1]-size)*pos[1],(shape[0]-size)*pos[0],(shape[1]-size)*pos[1]+size,(shape[0]-size)*pos[0]+size]

#print(pixel_to_angle_bars2(8))
#tta = time_to_angles(25,80,2,6,1000,40)
#print(pixel_to_angle_bars(tta[0],0),',',pixel_to_angle_bars(tta[1],0),',',pixel_to_angle_bars(tta[2],0))
#print(pixel_to_angle_bars(100,0),pixel_to_angle_bars2(10))
def repet_to_speed_bars2(rep,dir):
	#for new data: dir = 0
	dir = 0
	dir1 = 3.14/180*dir
	dir2 = real_angle(dir1,screens_ratio)
	s = speed_90bars30*30/rep
	s2 = real_dist(s,dir1,dir2,screens_ratio)
	stheta = s2/dist_screen_eyes*180/3.14
	return stheta

def repet_to_speed_bars(rep):
	stheta = 60/rep
	return stheta

def num_to_col_letters(num):
	letters = ''
	while num:
		mod = (num - 1) % 26
		letters += chr(mod + 65)
		num = (num - 1) // 26
	return ''.join(reversed(letters))

def std(frames):
	if frames is None:
		print('No frame for std')
		return None
	color = 0
	if type_of_image(frames[0]) == 'color':
		color = 1
	fm = np.zeros((frames.shape[1],frames.shape[2]))
	fs = np.zeros((frames.shape[1],frames.shape[2]))
	for t in range(0,frames.shape[0]):
		if color:
			gray = cv2.cvtColor(frames[t], cv2.COLOR_BGR2GRAY)
		else:
			gray = frames[t]
		fm+= gray/len(frames)
		fs+=(gray/len(frames))**2
	rep = np.sqrt(fs/len(frames)-(fm/len(frames))**2)
	return rep
def minmax(frames):
	return np.max(frames,0)-np.min(frames,0)

def mat_to_frame(mat1,scaling = True,bits = None):
	mat = mat1.copy()
	if mat is None:
		print('No matrix for mat_to_fame')
		return None
	if bits == None:
		bits = getbits(mat)
	mat = np.nan_to_num(mat,nan=0)
	if scaling:
		mat = scale(mat,bits = bits)
	return mat

def last_analysable_time(angle_lim,angles):
	idx = (np.abs(angles - angle_lim)).argmin()
	return idx

def rfmatrix(rfpos,ctpos,theta,length):
	theta = 3.14/180*theta
	dx = rfpos[0]-ctpos[0]
	dy = rfpos[1]-ctpos[1]
	d =  np.sqrt(dx*dx+dy*dy)
	a = np.arctan(dy/dx)
	if dx<0: a=3.14+a
	x = -d*np.cos(theta+a)
	return int(length/2+x)
	
def rfmatrix2(rfpos,theta,length):
	theta = 3.14/180*theta
	dx = rfpos[0]
	dy = rfpos[1]
	d =  np.sqrt(dx*dx+dy*dy)
	a = math.atan2(dy,dx)
	x = d*np.cos(a+theta)
	return int(length/2+x)
'''angle = 315
print(rfmatrix2([11,27],angle,1000))
print(rfmatrix2([10,20],angle,1000))
print(rfmatrix2([10,-20],angle,1000))
print(rfmatrix2([-10,-20],angle,1000))
'''
def rfmatrix_from_bar(time,trace,stim_name,length):
	rep = int(stim_name.split('bar')[-1])
	thick = int(stim_name.split('mov')[0])
	max_index1 = np.argmax(trace[0:int(length/(14400/rep)*len(trace))])
	print(time[max_index1])
	max_index2 = np.argmax(trace[int((length+thick)/(14400/rep)*len(trace)):int((2*length+thick)/(14400/rep)*len(trace))])
	max_index2 = max_index2 + int((length+thick)/(14400/rep)*len(trace))
	print(time[max_index2])
	pos1 = max_index1/len(trace)*(14400/rep)
	pos2 = max_index2/len(trace)*(14400/rep)
	return [pos1,pos2]

def norm(u):
	return np.sqrt(sum(u**2)) 
def proj(u,v):
	return (np.dot(u, v)/norm(v)**2)*v 
def projplane(u,P):
	pu = np.zeros(len(u))
	Pbase = np.zeros(len(P))
	for i in range(0,len(P)):
		v = proj(u,P[i])
		pu+= v
		Pbase[i] = norm(v)*((np.dot(P[i], v)>=0)*2-1)
	return {'obase':pu,'Pbase':Pbase}
def turn90(v,u):
	return v-(np.dot(u,v)/norm(u)**2)*u
def rfbarspos1_to_rfpos(px,length,angle):
	if angle==90:
		return length-px-length//2
	else:
		return px-length//2
def rfbarspos2_to_rfpos(px,length,stim_name,angle):
	thick = int(stim_name.split('mov')[0])
	if angle==90:
		return length-(length-(px-(length+thick)))-length//2
	else:
		return length-(px-(length+thick))-length//2
def pos1_to_angle(px,length):
	rep= -(px/length*150-75)
	return int(100*rep)/100
def pos2_to_angle(px,stim_name,length):
	thick = int(stim_name.split('mov')[0])
	rep= pos1_to_angle(length-(px-(length+thick)),length)
	return int(100*rep)/100

def parse_rf(file):
	pos = []
	color = []
	if (file is None):
		return {'pos':pos,'color':color}
	with open(file, 'r') as f:
		data = f.readlines()
	for line in data:
		chars = line.split()
		chars = np.asarray(chars)
		pos.append([float(chars[0]),float(chars[1]),float(chars[2]),float(chars[3])])
		color.append(chars[4])
	return {'pos':pos,'color':color}
def save_trace_xlsx(time,trace,path,name,sheet,recording):
	#check if excel file exists
	if not os.path.exists(path+"/"+name+".xlsx"):
		workbook = xlsxwriter.Workbook(path+"/"+name+".xlsx")
		worksheet = workbook.add_worksheet(sheet)
		workbook.close()
		
	workbook = openpyxl.load_workbook(path+"/"+name+".xlsx")
	#check if sheet exists
	if not sheet in workbook.sheetnames:
		worksheet = workbook.create_sheet(sheet)
	else:
		worksheet = workbook[sheet]
	#create time column if nothing in first cell
	if worksheet['A1'].value != 'time':
		worksheet.insert_cols(1)
		worksheet['A1'] = 'time'
		for row in range(0,len(time)):
			worksheet['A'+str(row+2)]=time[row]
	col = 2
	while worksheet[num_to_col_letters(col)+'1'].value!=recording and worksheet[num_to_col_letters(col)+'1'].value is not None:
		col += 1
	worksheet[num_to_col_letters(col)+'1'] = recording
	for row in range(1,len(trace)+1):
		worksheet[num_to_col_letters(col)+str(row+1)] = trace[row-1]
	workbook.save(path+"/"+name+".xlsx")
	workbook.close()

def save_trace_xls(time,trace,path,name,sheet,recording):
	#check if excel file exists
	if not os.path.exists(path+"/"+name+".xls"):
		workbook = xlsxwriter.Workbook(path+"/"+name+".xls")
		worksheet = workbook.add_worksheet(sheet)
		workbook.close()
	
	w = copy(path+"/"+name+".xls")
	sheets = w._Workbook__worksheets
	sheet_names = []
	for s in sheet:
		sheet_names.append(s.get_name())
	#check if sheet exists
	if not sheet in w.sheetnames:
		worksheet = w.create_sheet(sheet)
	else:
		worksheet = w.get_sheet(sheet)
	#create time column if nothing in first cell
	if worksheet['A1'].value != 'time':
		worksheet.insert_cols(1)
		worksheet['A1'] = 'time'
		for row in range(0,len(time)):
			worksheet['A'+str(row+2)]=time[row]
	col = 2
	while worksheet[num_to_col_letters(col)+'1'].value!=recording and worksheet[num_to_col_letters(col)+'1'].value is not None:
		col += 1
	worksheet[num_to_col_letters(col)+'1'] = recording
	for row in range(1,len(trace)+1):
		worksheet[num_to_col_letters(col)+str(row+1)] = trace[row-1]
	w.save(path+"/"+name+".xls")
	workbook.close()

def add_xlsx2(path,name,sheet,data,save=True):
	#check if excel file exists
	if not os.path.exists(path+"/"+name+".xlsx"):
		workbook = xlsxwriter.Workbook(path+"/"+name+".xlsx")
		worksheet = workbook.add_worksheet(sheet)
		workbook.close()
		
	workbook = xw.Book(path+"/"+name+".xlsx")
	worksheet = workbook.sheets[sheet]
	row = 1
	while worksheet.range('A'+str(row)).value is not None:
		row += 1
	col = 1
	for i in data:
		worksheet[num_to_col_letters(col)+str(row)].value = i
		col += 1
	if save:
		workbook.save(path+"/"+name+".xlsx")

def add_xlsx3(path,name,sheet,data,save=True,disp='row'):
	#check if excel file exists
	if not os.path.exists(path+"/"+name+".xlsx"):
		workbook = xlsxwriter.Workbook(path+"/"+name+".xlsx")
		worksheet = workbook.add_worksheet(sheet)
		workbook.close()
	workbook = xw.Book(path+"/"+name+".xlsx")
	#check if sheet exists
	if not sheet in [i.name for i in workbook.sheets]:
		worksheet = workbook.sheets.add(sheet)
	else:
		worksheet = workbook.sheets[sheet]
	if disp == 'row':
		row = 1
		while worksheet.range('A'+str(row)).value is not None:
			row += 1
		if len(np.asarray(data).shape)==1:
			data = [data]
		for datax in data:
			col = 1
			for i in datax:
				worksheet[num_to_col_letters(col)+str(row)].value = i
				col += 1
			row += 1
	elif disp == 'col':
		col = 1
		while worksheet.range(num_to_col_letters(col)+'1').value is not None:
			col += 1
		if len(np.asarray(data).shape)==1:
			data = [data]
		for datax in data:
			row = 1
			for i in datax:
				worksheet[num_to_col_letters(col)+str(row)].value = i
				row += 1
			col += 1
	else:
		print('disp argument doesnt exist, only col or row accepted you twat')
	if save:
		workbook.save(path+"/"+name+".xlsx")
def read_xlsx3(path,sheet):
	workbook = xw.Book(path)
	#check if sheet exists
	if not sheet in [i.name for i in workbook.sheets]:
		print('sheet doesnt exist in workbook ',path)
	else:
		worksheet = workbook.sheets[sheet]
	row = 1
	while worksheet.range('A'+str(row)).value is not None:
		row += 1
	col = 1
	while worksheet.range(num_to_col_letters(col)+'1').value is not None:
		col += 1
	return worksheet.range('A1:'+num_to_col_letters(col-1)+str(row-1)).value

	
def add_xlsx(path,name,sheet,data):
	#check if excel file exists
	if not os.path.exists(path+"/"+name+".xlsx"):
		workbook = xlsxwriter.Workbook(path+"/"+name+".xlsx")
		worksheet = workbook.add_worksheet(sheet)
		workbook.close()
		
	workbook = openpyxl.load_workbook(path+"/"+name+".xlsx")
	#check if sheet exists
	if not sheet in workbook.sheetnames:
		worksheet = workbook.create_sheet(sheet)
	else:
		worksheet = workbook[sheet]
	row = 1
	while worksheet['A'+str(row)].value is not None:
		row += 1
	col = 1
	for i in data:
		worksheet[num_to_col_letters(col)+str(row)] = i
		col += 1
	workbook.save(path+"/"+name+".xlsx")
	workbook.close()



times_videos_path = '//smallbrains-nas2/keivan/Data/Track ball/times_videos.txt'

colorsNames = ['maroon','green','navy','olive','teal','purple','gray','black','maroon','green','navy','olive','teal','purple','gray','black'] 	
colorsRGB = [(128,0,0),(0,128,0),(0,0,128),(128,128,0),(0,128,128),(128,0,128),(128,128,128),(0,0,0),(128,0,0),(0,128,0),(0,0,128),(128,128,0),(0,128,128),(128,0,128),(128,128,128),(0,0,0)]
colorsRGB2 = [(0,0,128),(0,128,0),(128,0,0),(0,128,128),(128,128,0),(128,0,128),(128,128,128),(0,0,0),(0,0,128),(0,128,0),(128,0,0),(0,128,128),(128,128,0),(128,0,128),(128,128,128),(0,0,0)]
def colorsRGB_unlim(i):
	return colorsRGB[i % 8]
def color_names(i):
	return colorsNames[i % 8]
names_to_RGB = {
	'maroon':(128,0,0),
	'green':(0,128,0),
	'navy':(0,0,128),
	'olive':(128,128,0),
	'teal':(0,128,128),
	'purple':(128,0,128),
	'gray':(128,128,128),
	'black':(0,0,0)
	}
	
